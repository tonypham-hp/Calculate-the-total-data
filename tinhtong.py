import os
import pandas as pd
import re
from openpyxl import load_workbook, Workbook


#ROOT_PATH = r"\\10.147.32.1\MA_Div\Data_Link"
#ROOT_PATH = r"Documents"
ROOT_PATH = r"c:\Users\vnintern09\Documents"
 
#cac nam FY
#ACTUAL_ROOT = os.path.join(ROOT_PATH, "CostDX-BM-TH-ACT", "XUẤT KHO FY25")
#tong ket
#BUDGET_PATH = os.path.join(ROOT_PATH, "BUDGET FY25")
#luu tam
#TEMP_PATH = os.path.join(ROOT_PATH, "Temp_Result")

try:
    # Khởi tạo đường dẫn
    ACTUAL_ROOT = os.path.join(ROOT_PATH, "CostDX-BM-TH-ACT", "XUẤT KHO FY25")
    BUDGET_PATH = os.path.join(ROOT_PATH, "BUDGET FY25")
    TEMP_PATH   = os.path.join(ROOT_PATH, "Temp_Result") 

    # Kiểm tra tồn tại từng thư mục
    if not os.path.exists(ACTUAL_ROOT):
        raise FileNotFoundError(f"Không tìm thấy thư mục Actual: {ACTUAL_ROOT}")
    if not os.path.exists(BUDGET_PATH):
        raise FileNotFoundError(f"Không tìm thấy thư mục Budget: {BUDGET_PATH}")
    # Temp có thể chưa tồn tại → tự tạo
    if not os.path.exists(TEMP_PATH):
        os.makedirs(TEMP_PATH, exist_ok=True)

except Exception as e:
    print(f"Lỗi cấu hình đường dẫn dữ liệu: {e}")


#danh sach phan loai san pham
#CATEGORIES = ["BM", "PM", "CM", "TH", "Tiêu hao", "TIÊU HAO"]
CATEGORIES = ["BM", "PM", "CM", "TH"]
#danh sach cac thang nam tai chinh
MONTH_MAPPING = {
    "Tháng 04.2025": "APR",
    "Tháng 05.2025": "MAY",
    "Tháng 06.2025": "JUN",
    "Tháng 07.2025": "JUL",
    "Tháng 08.2025": "AUG",
    "Tháng 09.2025": "SEP",
    "Tháng 10.2025": "OCT",
    "Tháng 11.2025": "NOV",
    "Tháng 12.2025": "DEC",
    "Tháng 01.2026": "JAN",
    "Tháng 02.2026": "FEB",
    "Tháng 03.2026": "MAR"
}

def normalize_category(raw_value):
    """'Tiêu hao' → 'TH'"""
    if pd.isna(raw_value):
        return None
    val = str(raw_value).strip().upper()
    if "TIÊU HAO" in val:
        return "TH"
    return val if val in CATEGORIES else None
#-----------------------------
# P1: Tinh tong + ghi file tam
#-----------------------------
                  
def process_actual(fiscal_year = "FY25"):
    print("phan 1 bat dau")
    print(f"ACTUAL_ROOT: {ACTUAL_ROOT}")
    if not os.path.exists(ACTUAL_ROOT):
        #khong tim thay thu muc
        print(f"khong co {ACTUAL_ROOT}")
        return
    
    os.makedirs(TEMP_PATH, exist_ok=True)

    #danh sách theo thứ tự năm tài chính
    months = sorted (
        [f for f in os.listdir(ACTUAL_ROOT) if f in MONTH_MAPPING], 
        key=lambda x: list(MONTH_MAPPING.keys()).index(x)
    )
    print(f"tim thay {len(months)} thang: {months}")

    #quét tất cả các tháng từ 312
    all_products = []
    print("\n=== quet tat ca ma line theo thu tu ===")
    
    for month_idx, month_folder in enumerate(months, 1):
        print(f"\n[{month_idx}/{len(months)}] quet thang: {month_folder}")

        month_path = os.path.join(ACTUAL_ROOT, month_folder)
        #tìm theo file "theo dõi xuất kho"

        excel_file = None
        for f in os.listdir(month_path):
            #dieu kien file .xlsx co chu "theo doi xuat kho" trong name
            if f.endswith(".xlsx") and "Theo dõi Xuất kho" in f:
                excel_file = f
                break
        if not excel_file:
            print(f"khong co file theo doi xuat kho")
            continue
        print(f" file: {excel_file}")
        file_path = os.path.join(month_path, excel_file)

        #tim sheet chứa XUẤT T
        excel = pd.ExcelFile(file_path)
        sheet_name = None
        for sheet in excel.sheet_names:
            #sheet co chua "XUAT T"
            if "XUẤT T" in sheet.upper():
                sheet_name = sheet
                break
        if not sheet_name:
            print(f"khong co sheet xuat T")
            continue
        print(f"sheet: {sheet_name}")

        try:
            # Đọc dữ liệu cột M,J,W
            df = pd.read_excel(file_path, sheet_name=sheet_name, usecols="J,M,W", skiprows=1)
            df.columns = ["Tổng tiền","Mã Line","Phân loại"]
            print(f"  Raw data shape: {df.shape}")
            #bo header + NaN
            df = df.dropna(subset=["Mã Line"]).reset_index(drop=True)
            print(f"  Sau drop NaN: {df.shape}")
            print(f"  Mẫu RAW Mã Line: {df['Mã Line'].head(3).tolist()}")
            print(f"  Mẫu RAW Tổng tiền: {df['Tổng tiền'].head(3).tolist()}")
            

            df["Mã Line"] = df["Mã Line"].astype(str).str.strip()
            # Lọc bỏ mã 312 và NaN
            month_codes = [code for code in df["Mã Line"].unique() if code != "312" and code != "Mã Line" and code != "nan"]
            
            # DEBUG: In + them tất cả mã Line unique của tháng này
            for code in month_codes:
                if code not in all_products:
                    all_products.append(code)
                    print(f"    + NEW CODE: {code}")
            
            print(f"  Tháng này: {len(month_codes)} mã")
            print(f"  Tổng cộng: {len(all_products)} mã") 
                        

        except Exception as e:
            print(f" loi doc file {e}")
            continue

    print(f"\n=== TỔNG KẾT CUỐI: {len(all_products)} MÃ ===")
    all_products = sorted(all_products, key=lambda x: int(x) if x.isdigit() else 0)
    print("Danh sách mã (đã sort):", all_products[:10], "...")
    

    #tinh tong cho tung ma 
    print("\n=== TÍNH BM/PM/CM/TH CHO TỪNG MÃ ===")
    yearly_data = {}

    for product_idx, product in enumerate(sorted(all_products), 1):
        print(f"\n[{product_idx}/{len(all_products)}] Tinh Ma {product}")
        yearly_data[product] = {}

        for month_idx, month_folder in enumerate(months, 1):
            print(f"  [{month_idx}/{len(months)}] {month_folder}", end=" -> ")
            month_path = os.path.join(ACTUAL_ROOT, month_folder)
            #excel_files = [f for f in os.listdir(month_path) if re.match(r"1\.\s*Theo\s*dõi\s*Xuất\s*kho\s*T\d+\.xlsx", f)]
            excel_file = None
            for f in os.listdir(month_path):
                if f.endswith(".xlsx") and "Theo dõi Xuất kho" in f:
                    excel_file = f 
                    break
            if not excel_file:
                yearly_data[product][month_folder] = {c: 0 for c in CATEGORIES}
                continue

            file_path = os.path.join(month_path, excel_file)

            excel = pd.ExcelFile(file_path)
            sheet_name = None
            for sheet in excel.sheet_names:
                if "XUẤT T" in sheet.upper():
                    sheet_name = sheet
                    break
            if not sheet_name:
                print("khong co sheet")
                yearly_data[product][month_folder] = {c: 0 for c in CATEGORIES}
                continue

            try:
                #dung thu tu
                df = pd.read_excel(file_path, sheet_name=sheet_name, usecols="J,M,W")
                df.columns = ["Tổng tiền","Mã Line","Phân loại"]
                
                # CHỈ LẤY DÒNG CỦA MÃ NÀY
                df["Mã Line"] = df["Mã Line"].astype(str).str.strip()
                df_product = df[df["Mã Line"] == product]
                print("===="*10)
                print(df_product)
                print("===="*10)

                print(f"({len(df_product)} dòng)", end=" -> ")

                df_product = df_product[df_product["Mã Line"] != 312]
                df_product["CatKey"] = df_product["Phân loại"].apply(normalize_category)
                df_product = df_product.dropna(subset=["CatKey"])
                
                print(f"({len(df_product)} dòng hợp lệ)", end=" -> ")
                # Tính tổng riêng từng phân loại
                month_totals = {c: 0 for c in CATEGORIES}
                if not df_product.empty:
                    grouped = df_product.groupby("CatKey")["Tổng tiền"].sum()
                    for cat in CATEGORIES:
                        if cat in grouped.index:
                            month_totals[cat] = float(grouped[cat])
                    print(f"BM={month_totals['BM']:,.0f} PM={month_totals['PM']:,.0f} "
                      f"CM={month_totals['CM']:,.0f} TH={month_totals['TH']:,.0f}")
                else:
                    print("KHONG CO PHAN LOAI HOP LE")
                yearly_data[product][month_folder] = month_totals
                
            except Exception as e:
                print(f"Loi: {e}")
                yearly_data[product][month_folder] = {c: 0 for c in CATEGORIES}

        print(f"\n  GHI FILE row.{product}.{fiscal_year}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = f"Dữ liệu {fiscal_year}"
        #header: Month | BM | PM | CM | TH
        ws.append(["Month"] + CATEGORIES)
        
        for month_folder in months:
            month_data = yearly_data[product][month_folder]
            #tao dong: [Thang, bm_tong, pm_tong, cm_tong, th_tong]
            row_data = [month_folder] + [month_data[c] for c in CATEGORIES]
            ws.append(row_data)

        #luu file row.<ma>.FY25.xlsx
        filename = os.path.join(TEMP_PATH, f"row.{product}.{fiscal_year}.xlsx")
        wb.save(filename)
        print(f"  OK Luu row.{product}.FY25.xlsx")
    print("phan 1 hoan thanh")

#-------------------------------------
# P2: cap nhat du lieu vao file budget
#-------------------------------------

def update_budget(fiscal_year="FY25"):
    print("\n=== PHẦN 2: CẬP NHẬT BUDGET ===")
    
    # Kiểm tra thư mục TEMP
    if not os.path.exists(TEMP_PATH):
        print(f"Không tìm thấy {TEMP_PATH}")
        return
    
    # Lấy tất cả file row.*.FY25.xlsx
    temp_files = [f for f in os.listdir(TEMP_PATH) if f.startswith("row.") and f.endswith(f".{fiscal_year}.xlsx")]
    print(f"Tìm thấy {len(temp_files)} file row trong Temp:")
    for f in temp_files[:3]:  # In 3 file đầu
        print(f"  - {f}")
    if len(temp_files) > 3:
        print(f"  ... và {len(temp_files)-3} file nữa")
    
    if not temp_files:
        print("Không có file row nào để update!")
        return
    
    # Kiểm tra thư mục BUDGET
    if not os.path.exists(BUDGET_PATH):
        print(f"Không tìm thấy {BUDGET_PATH}")
        return
    
    update_count = 0
    for temp_idx, temp_file in enumerate(temp_files, 1):
        # Lấy mã sản phẩm từ tên file: row.461.FY25.xlsx → "461"
        product = temp_file.split(".")[1]
        print(f"\n[{temp_idx}/{len(temp_files)}] Cập nhật MÃ {product}: {temp_file}")
        
        temp_path = os.path.join(TEMP_PATH, temp_file)
        df_temp = pd.read_excel(temp_path)
        print(f"Đọc file row OK: {len(df_temp)} dòng")
        
        # Tìm file budget tương ứng: 461.25.xlsx
        budget_file = os.path.join(BUDGET_PATH, f"{product}.25.xlsx")
        if not os.path.exists(budget_file):
            print(f" Không có budget file: {budget_file}")
            continue
        
        print(f"  Mở budget file: {budget_file}")
        wb = load_workbook(budget_file)
        
        # Với mỗi THÁNG trong file row
        for idx, row in df_temp.iterrows():
            month_folder = str(row["Month"])  # "Tháng 04.2025"
            budget_sheet_name = MONTH_MAPPING.get(month_folder)  # "APR"
            
            if not budget_sheet_name or budget_sheet_name not in wb.sheetnames:
                print(f"  Skip tháng {month_folder} (sheet {budget_sheet_name} không tồn tại)")
                continue
            
            ws = wb[budget_sheet_name]
            print(f" Update sheet '{budget_sheet_name}' ({month_folder})")
            
            # Duyệt TẤT CẢ dòng trong sheet budget
            for r in range(2, ws.max_row + 1):  # Bắt đầu từ dòng 2 (bỏ header)
                # Cột F: Kiểm tra "xuất kho"
                item_cell = ws[f"F{r}"]
                item = str(item_cell.value).strip().lower() if item_cell.value else ""
                
                if item == "xuất kho":
                    # Cột N: Phân loại (BM, PM, CM, TH, Tiêu hao...)
                    remark_cell = ws[f"N{r}"]
                    remark_raw = str(remark_cell.value).strip() if remark_cell.value else ""
                    
                    # Chuẩn hóa phân loại
                    cat_key = normalize_category(remark_raw)
                    if cat_key and cat_key in CATEGORIES:
                        # Lấy giá trị từ file row tương ứng
                        value = float(row[cat_key])
                        
                        # Ghi vào cột I
                        ws[f"I{r}"].value = value
                        update_count += 1
                        print(f"Dòng {r}: {remark_raw} → {cat_key} = {value:,.0f}")
                    else:
                        print(f"Dòng {r}: Phân loại '{remark_raw}' không hợp lệ")
        
        # Lưu file budget
        wb.save(budget_file)
        print(f" Lưu budget file: {product}.25.xlsx")
    
    print(f"\n=== PHẦN 2 HOÀN THÀNH ===")
    print(f"Tổng cộng: {update_count} ô dữ liệu đã update")
    print(f"{len(temp_files)} file budget đã được cập nhật")

def main():
    process_actual("FY25")
    update_budget("FY25")

if __name__ == "__main__":
    main()       
